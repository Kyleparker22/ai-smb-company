#!/usr/bin/env python3
"""Edit the financial model's input assumptions from HQ — with an honest pending state.

The workbook's ~6,800 formulas can only be evaluated by a spreadsheet engine, and the
machine serving HQ has none. So an edit here does two things and stops:

  1. writes the new value into the workbook's INPUT cell (never a formula)
  2. records it as PENDING RECALCULATION

Until `runtime/finance_model_recalc.py` runs, HQ keeps showing the last computed
figures and states plainly that they do not include the pending edits. It must never
show a number that looks like the consequence of a change it has not actually
computed — that is the whole reason this is a two-step flow instead of one.

Cells are resolved BY LABEL, never by row number: rows in this workbook have moved
four times, and a registry of coordinates would silently start writing into the wrong
assumption. Every write is verified to land on a literal, not a formula.

Run:
    python3 runtime/finance_model_edit.py --list
    python3 runtime/finance_model_edit.py --set cogsPerClient=250 --by the Founder
    python3 runtime/finance_model_edit.py --pending
"""
import json, os, sys, argparse, datetime

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.environ.get("YOURCO_DATA_ROOT") or os.path.dirname(HERE)
XLSX = os.path.join(ROOT, "finance", "yourco-financial-model.xlsx")
PENDING = os.path.join(ROOT, "dashboard", "finance_pending.json")
SCEN_COLS = {"Conservative": 3, "Target": 4, "Aggressive": 5}

# key -> (label on the Assumptions sheet, column, kind, low, high, human label)
# `column` 2 = the single-value column B; a scenario name means that case's column.
EDITABLE = {
    "cogsPerClient":      ("COGS — absorbed model/voice/hosting per client", 2, "money", 0, 5000,
                           "COGS / client / month"),
    "advisorFullyLoaded": ("Advisor — fully loaded / month", 2, "money", 0, 60000,
                           "Advisor, fully loaded / month"),
    "clientsPerAdvisor":  ("Clients one Advisor can carry", 2, "int", 1, 200,
                           "Clients one Advisor carries"),
    "principalCapacity":  ("Clients the three principals carry between them", 2, "int", 0, 500,
                           "Principals' capacity (before the glide)"),
    "capacityHoldMonth":  ("Capacity holds flat until month", 2, "int", 0, 36,
                           "Capacity holds flat until month"),
    "capacityZeroMonth":  ("Capacity reaches zero at month", 2, "int", 1, 36,
                           "Capacity reaches zero at month"),
    "principalSalaryFounder":  ("the Founder — principal salary", 2, "money", 0, 500000, "the Founder — annual salary"),
    "principalSalaryPartnerB": ("Partner B — principal salary", 2, "money", 0, 500000, "Partner B — annual salary"),
    "principalSalaryMike":  ("Mike — principal salary", 2, "money", 0, 500000, "Mike — annual salary"),
    "commissionEach":     ("Commission — % of revenue, each", 2, "pct", 0, 0.5,
                           "Commission — % of revenue, each"),
    "distributionPctMRR": ("Operating distribution — % of MRR", 2, "pct", 0, 0.5,
                           "Operating distribution — % of MRR"),
    "taxRate":            ("Tax distribution — assumed effective rate", 2, "pct", 0, 0.6,
                           "Tax distribution — assumed rate"),
    "onboardingCost":     ("Token + tool cost per new client", 2, "money", 0, 20000,
                           "Onboarding $ / client"),
    "onboardingHours":    ("Principal/operator hours per new client", 2, "num", 0, 500,
                           "Onboarding hours / client"),
    "cac":                ("CAC — cost to acquire one client", 2, "money", 0, 100000, "CAC"),
    "activeScenario":     ("Active scenario", 2, "int", 1, 3,
                           "Active scenario (1 Conservative · 2 Target · 3 Aggressive)"),
}
for case in SCEN_COLS:
    EDITABLE[f"conversion{case}"] = ("Audit → paying client conversion", case, "pct", 0.01, 1.0,
                                     f"{case} — audit→client conversion")
    EDITABLE[f"churn{case}"] = ("Annual logo churn", case, "pct", 0.0, 0.9,
                                f"{case} — annual logo churn")
    for y in (1, 2, 3):
        EDITABLE[f"targetY{y}{case}"] = (f"Active clients — end of Year {y}", case, "int", 0, 5000,
                                         f"{case} — clients target, end of Year {y}")


def _find(ws, text):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and text in v:
            return r
    return None


def _resolve(wb, key):
    label, col, *_ = EDITABLE[key]
    ws = wb["Assumptions"]
    r = _find(ws, label)
    if not r:
        raise LookupError(f"'{label}' not found on the Assumptions sheet — the workbook changed shape")
    c = SCEN_COLS[col] if isinstance(col, str) else col
    return ws, r, c


def read_current():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX)
    out = {}
    for key in EDITABLE:
        try:
            ws, r, c = _resolve(wb, key)
            v = ws.cell(r, c).value
            out[key] = None if isinstance(v, str) and v.startswith("=") else v
        except LookupError:
            out[key] = None
    return out


def load_pending():
    if not os.path.exists(PENDING):
        return {"edits": [], "since": None}
    try:
        return json.load(open(PENDING))
    except Exception:
        return {"edits": [], "since": None}


def apply(edits, by="hq"):
    """Write input values into the workbook and mark the model pending recalculation."""
    import openpyxl
    if not edits:
        return {"ok": False, "error": "no edits given"}
    unknown = [k for k in edits if k not in EDITABLE]
    if unknown:
        return {"ok": False, "error": f"not editable: {', '.join(unknown)}"}

    wb = openpyxl.load_workbook(XLSX)
    applied, problems = [], []
    for key, raw in edits.items():
        label, col, kind, lo, hi, human = EDITABLE[key]
        try:
            val = float(raw)
        except (TypeError, ValueError):
            problems.append(f"{human}: {raw!r} is not a number")
            continue
        if kind == "int":
            val = int(round(val))
        if not (lo <= val <= hi):
            problems.append(f"{human}: {val} is outside the allowed range {lo}–{hi}")
            continue
        try:
            ws, r, c = _resolve(wb, key)
        except LookupError as e:
            problems.append(str(e))
            continue
        cur = ws.cell(r, c).value
        # Never write over a formula — that would delete a derivation and leave a
        # constant that silently stops responding to the rest of the model.
        if isinstance(cur, str) and cur.startswith("="):
            problems.append(f"{human}: target cell holds a formula, refusing to overwrite it")
            continue
        if cur == val:
            continue
        ws.cell(r, c).value = val
        applied.append({"key": key, "label": human, "cell": f"{ws.cell(r, c).coordinate}",
                        "from": cur, "to": val})

    if problems and not applied:
        return {"ok": False, "error": "; ".join(problems)}
    if not applied:
        return {"ok": True, "applied": [], "note": "no change — values already match"}

    wb.save(XLSX)
    p = load_pending()
    p["since"] = p.get("since") or datetime.datetime.now().isoformat(timespec="seconds")
    for a in applied:
        a["at"] = datetime.datetime.now().isoformat(timespec="seconds")
        a["by"] = by
        p["edits"].append(a)
    p["note"] = ("These values are IN the workbook but the workbook has not been recalculated, so "
                 "every computed figure HQ is showing predates them.")
    p["fix"] = "python3 runtime/finance_model_recalc.py"
    with open(PENDING, "w") as f:
        json.dump(p, f, indent=2, ensure_ascii=False)
    return {"ok": True, "applied": applied, "problems": problems, "pending": len(p["edits"])}


def clear_pending():
    if os.path.exists(PENDING):
        os.remove(PENDING)


def main():
    ap = argparse.ArgumentParser(description="Edit the financial model's input assumptions.")
    ap.add_argument("--list", action="store_true")
    ap.add_argument("--set", action="append", metavar="KEY=VALUE", default=[])
    ap.add_argument("--by", default="cli")
    ap.add_argument("--pending", action="store_true")
    a = ap.parse_args()

    if a.list:
        cur = read_current()
        print(f"{len(EDITABLE)} editable inputs:\n")
        for k, (label, col, kind, lo, hi, human) in EDITABLE.items():
            print(f"  {k:<24} {str(cur.get(k)):<12} {human}  [{kind} {lo}–{hi}]")
        return 0
    if a.pending:
        p = load_pending()
        if not p["edits"]:
            print("nothing pending — HQ's figures reflect the workbook")
            return 0
        print(f"{len(p['edits'])} pending edit(s) since {p['since']} — NOT yet recalculated:")
        for e in p["edits"]:
            print(f"  {e['label']}: {e['from']} → {e['to']}  ({e['cell']}, by {e['by']})")
        print(f"\n  {p['fix']}")
        return 1
    if a.set:
        edits = {}
        for pair in a.set:
            if "=" not in pair:
                print(f"bad --set {pair!r}, expected KEY=VALUE", file=sys.stderr)
                return 2
            k, v = pair.split("=", 1)
            edits[k.strip()] = v.strip()
        r = apply(edits, by=a.by)
        if not r.get("ok"):
            print("refused: " + r["error"], file=sys.stderr)
            return 2
        for e in r.get("applied", []):
            print(f"set {e['label']}: {e['from']} → {e['to']}  ({e['cell']})")
        for pb in r.get("problems", []):
            print(f"skipped — {pb}", file=sys.stderr)
        if r.get("applied"):
            print(f"\n⚠ PENDING RECALCULATION ({r['pending']} edit(s)). "
                  f"HQ's figures still predate this.\n  python3 runtime/finance_model_recalc.py")
        return 0
    ap.print_help()
    return 0


if __name__ == "__main__":
    sys.exit(main())
