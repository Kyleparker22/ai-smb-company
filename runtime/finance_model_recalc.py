#!/usr/bin/env python3
"""Recalculate the financial model, then clear the pending state and re-sync HQ.

This is the second half of HQ editing. `finance_model_edit.py` writes an input value
and marks the model pending; nothing downstream is true until a spreadsheet engine has
re-evaluated the ~6,800 formulas. This does that, checks the result, and only then
tells HQ the figures are current again.

Needs Excel on macOS. The VPS has neither Excel nor LibreOffice, so this is a
the Founder's-Mac operation by design — and it says so rather than failing obscurely.

Run:
    python3 runtime/finance_model_recalc.py
    python3 runtime/finance_model_recalc.py --dry-run
"""
import json, os, sys, subprocess, shutil

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.environ.get("YOURCO_DATA_ROOT") or os.path.dirname(HERE)
XLSX = os.path.join(ROOT, "finance", "yourco-financial-model.xlsx")
sys.path.insert(0, HERE)
import finance_model_edit as fme


def _osa(script, timeout=180):
    return subprocess.run(["osascript", "-e", script], capture_output=True, text=True, timeout=timeout)


def excel_available():
    if sys.platform != "darwin":
        return False, "not macOS"
    if not os.path.isdir("/Applications/Microsoft Excel.app"):
        return False, "Microsoft Excel is not installed"
    return True, None


def recalc():
    name = os.path.basename(XLSX)
    r = _osa(f'tell application "Microsoft Excel" to open POSIX file "{XLSX}"')
    if r.returncode != 0:
        return False, f"could not open the workbook in Excel: {r.stderr.strip()}"
    chk = _osa('tell application "Microsoft Excel" to return name of workbooks')
    if name not in (chk.stdout or ""):
        return False, ("Excel did not open the workbook — it is usually showing a modal dialog. "
                       "Bring Excel to the front, dismiss it, and re-run.")
    r = _osa(f'''tell application "Microsoft Excel"
    calculate
    delay 6
    save workbook "{name}"
    delay 2
    close workbook "{name}" saving no
end tell''')
    if r.returncode != 0:
        return False, f"recalculation failed: {r.stderr.strip()}"
    return True, None


def verify():
    """Zero formula errors, and every formula actually carries a computed value."""
    import openpyxl
    wf = openpyxl.load_workbook(XLSX)
    wv = openpyxl.load_workbook(XLSX, data_only=True)
    total = errors = uncached = 0
    first = []
    for ws in wf.worksheets:
        wsv = wv[ws.title]
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    total += 1
                    v = wsv[c.coordinate].value
                    if v is None:
                        uncached += 1
                    elif isinstance(v, str) and v.startswith("#"):
                        errors += 1
                        if len(first) < 5:
                            first.append(f"{ws.title}!{c.coordinate} {v}")
    return {"formulas": total, "errors": errors, "uncached": uncached, "sample": first}


def main():
    dry = "--dry-run" in sys.argv
    p = fme.load_pending()
    if not p["edits"]:
        print("nothing pending — the workbook and HQ already agree.")
        return 0

    print(f"{len(p['edits'])} pending edit(s) since {p['since']}:")
    for e in p["edits"]:
        print(f"  {e['label']}: {e['from']} → {e['to']}   ({e['cell']}, by {e['by']})")
    if dry:
        print("\n--dry-run: nothing recalculated.")
        return 0

    ok, why = excel_available()
    if not ok:
        print(f"\ncannot recalculate here — {why}.\n"
              f"This step needs a spreadsheet engine, and only the Founder's Mac has one.\n"
              f"Manual equivalent: open finance/yourco-financial-model.xlsx, let it calculate, save,\n"
              f"then run:  python3 runtime/finance_model_sync.py\n"
              f"and clear the pending flag:  python3 -c \"import sys;sys.path.insert(0,'runtime');"
              f"import finance_model_edit as f;f.clear_pending()\"", file=sys.stderr)
        return 2

    print("\nrecalculating in Excel…")
    ok, why = recalc()
    if not ok:
        print(f"failed: {why}", file=sys.stderr)
        return 2

    v = verify()
    print(f"  {v['formulas']} formulas · {v['errors']} errors · {v['uncached']} uncached")
    if v["errors"]:
        for s in v["sample"]:
            print(f"    {s}")
        print("\nRefusing to clear the pending state: the workbook has formula errors.\n"
              "The edit is still in the file — fix the errors, then re-run.", file=sys.stderr)
        return 3

    r = subprocess.run([sys.executable, os.path.join(HERE, "finance_model_sync.py")],
                       capture_output=True, text=True, cwd=ROOT)
    print(r.stdout.strip() or r.stderr.strip())
    if r.returncode != 0:
        print("sync failed — pending state left in place.", file=sys.stderr)
        return 4

    fme.clear_pending()
    print("\npending cleared — HQ now shows figures computed from the edits.")
    print("Commit the workbook AND dashboard/finance_model.json together.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
