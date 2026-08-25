#!/usr/bin/env python3
"""Sync the financial model into HQ — one direction, on purpose.

`finance/yourco-financial-model.xlsx` is canonical. It holds ~6,800 formulas off a
single Assumptions sheet, and only a spreadsheet engine can evaluate them. HQ reads
what this script extracts; it does not compute the model and it does not edit it.

WHY IT IS ONE DIRECTION. Writing an assumption back from HQ is easy; making the
other 6,800 cells correct afterwards is not — that needs Excel or LibreOffice, and
the VPS that serves HQ has neither. A dashboard that let you change a number and
then showed you figures computed from the OLD number would be worse than one that
refuses. So: edit the workbook, run this, commit both. The staleness check below is
what makes that safe — HQ can always tell whether what it is showing still matches
the file, and says so instead of quietly serving old numbers.

Run:
    python3 runtime/finance_model_sync.py            # sync
    python3 runtime/finance_model_sync.py --check    # exit 1 if HQ is stale
"""
import json, os, sys, hashlib, datetime, subprocess

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.environ.get("YOURCO_DATA_ROOT") or os.path.dirname(HERE)
XLSX = os.path.join(ROOT, "finance", "yourco-financial-model.xlsx")
OUT = os.path.join(ROOT, "dashboard", "finance_model.json")


def digest(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _commit():
    try:
        return subprocess.run(["git", "log", "-1", "--format=%h %ad", "--date=short", "--",
                               "finance/yourco-financial-model.xlsx"],
                              cwd=ROOT, capture_output=True, text=True, timeout=10).stdout.strip()
    except Exception:
        return ""


def find(ws, text, lo=1, hi=None, exact=False):
    hi = hi or ws.max_row
    for r in range(lo, hi + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and ((v.strip() == text) if exact else (text in v)):
            return r
    return None


def extract():
    import openpyxl
    wv = openpyxl.load_workbook(XLSX, data_only=True)
    wf = openpyxl.load_workbook(XLSX)
    A, S, P, SE = wv["Assumptions"], wv["Scenarios"], wv["Monthly P&L"], wv["Scenario Engines"]
    SEf = wf["Scenario Engines"]

    # Every cached formula is None until the workbook has been recalculated in a
    # spreadsheet app. Extracting from an uncalculated file would publish blanks to
    # HQ as if they were zeros, so refuse instead.
    if S.cell(15, 3).value is None:
        raise RuntimeError(
            "the workbook has no computed values — open it in Excel, let it calculate, "
            "save, then re-run. (openpyxl writes formulas without results.)")

    cases = ["Conservative", "Target", "Aggressive"]
    starts = [r for r in range(1, SEf.max_row + 1)
              if isinstance(SEf.cell(r, 1).value, str) and SEf.cell(r, 1).value.endswith("CASE")]

    def annual(row, year, block_lo, block_hi):
        return sum(SE.cell(row, c).value or 0 for c in range(2 + 12 * (year - 1), 2 + 12 * year))

    scen = {}
    for i, st in enumerate(starts):
        end = starts[i + 1] - 1 if i + 1 < len(starts) else SEf.max_row
        rev = find(SEf, "TOTAL REVENUE", st, end)
        eb = find(SEf, "EBITDA", st, end)
        act = find(SEf, "ACTIVE CLIENTS", st, end)
        ppl = find(SEf, "Total people", st, end)
        adv = find(SEf, "Advisors required", st, end)
        yrs = []
        for y in (1, 2, 3):
            r_, e_ = annual(rev, y, st, end), annual(eb, y, st, end)
            yrs.append({
                "year": y,
                "clients": round(SE.cell(act, 1 + 12 * y).value or 0, 1),
                "revenue": round(r_),
                "ebitda": round(e_),
                "noiPct": round(e_ / r_, 4) if r_ else None,
                "people": int(SE.cell(ppl, 1 + 12 * y).value or 0),
                "advisors": int(SE.cell(adv, 1 + 12 * y).value or 0),
                "arr": round((S.cell(12 + y, 2 + i).value or 0)),
            })
        scen[cases[i]] = {
            "years": yrs,
            "peakCash": round(S.cell(20, 2 + i).value or 0),
            "breakevenMonth": S.cell(21, 2 + i).value,
            "cumCashM36": round(S.cell(23, 2 + i).value or 0),
            "grossMargin": round(S.cell(18, 2 + i).value or 0, 4),
            "conversion": S.cell(5, 2 + i).value,
            "churn": S.cell(6, 2 + i).value,
        }

    # Principal earnings tab: 3-year totals per case
    earn = {}
    if "Principal Earnings" in wv.sheetnames:
        E = wv["Principal Earnings"]
        Ef = wf["Principal Earnings"]
        for i, case in enumerate(cases):
            hdr = None
            for r in range(1, Ef.max_row + 1):
                if str(Ef.cell(r, 1).value).strip().upper() == case.upper():
                    hdr = r
                    break
            if hdr:
                tot = find(Ef, "3-year total", hdr, hdr + 8)
                if tot:
                    earn[case] = {who: round(E.cell(tot, 2 + j).value or 0)
                                  for j, who in enumerate(("the Founder", "Partner B", "Mike"))}

    def acell(label, col=2):
        r = find(wf["Assumptions"], label)
        return A.cell(r, col).value if r else None

    targets = {}
    tr = find(wf["Assumptions"], "Active clients — end of Year 1")
    if tr:
        for j, y in enumerate((1, 2, 3)):
            targets[f"Y{y}"] = {cases[k]: A.cell(tr + j, 3 + k).value for k in range(3)}

    return {
        "generated": datetime.datetime.now().isoformat(timespec="seconds"),
        "source": {
            "file": "finance/yourco-financial-model.xlsx",
            "sha256": digest(XLSX),
            "modified": datetime.datetime.fromtimestamp(os.path.getmtime(XLSX)).isoformat(timespec="seconds"),
            "lastCommit": _commit(),
        },
        "activeScenario": A["B7"].value,
        "horizon": {"start": P.cell(11, 2).value, "end": P.cell(11, 37).value},
        "scenarios": scen,
        "principalEarnings3yr": earn,
        "clientTargets": targets,
        "assumptions": {
            "blendedRetainerEarly": acell("Blended retainer / client / month — early"),
            "blendedRetainerMature": acell("Blended retainer / client / month — mature"),
            "cogsPerClient": acell("COGS — absorbed model/voice/hosting per client"),
            "advisorFullyLoaded": acell("Advisor — fully loaded / month"),
            "clientsPerAdvisor": acell("Clients one Advisor can carry"),
            "principalCapacity": acell("Clients the three principals carry between them"),
            "capacityZeroMonth": acell("Capacity reaches zero at month"),
            "principalSalary": acell("the Founder — principal salary"),
            "commissionEach": acell("Commission — % of revenue, each"),
            "distributionPctMRR": acell("Operating distribution — % of MRR"),
            "onboardingCost": acell("Token + tool cost per new client"),
            "onboardingHours": acell("Principal/operator hours per new client"),
        },
        "ownership": {"the Founder": 0.50, "Partner B": 0.35, "Mike": 0.15},
        "honesty": ("Every figure here is a PLAN on assumptions with no evidence behind them — the "
                    "company is pre-revenue with zero clients and zero audits delivered. HQ mirrors "
                    "the workbook; it does not compute or edit it."),
    }


def main():
    if not os.path.exists(XLSX):
        print(f"missing {XLSX}", file=sys.stderr)
        return 2
    live = digest(XLSX)

    if "--check" in sys.argv:
        if not os.path.exists(OUT):
            print("HQ has never been synced — run: python3 runtime/finance_model_sync.py", file=sys.stderr)
            return 1
        synced = (json.load(open(OUT)).get("source") or {}).get("sha256")
        if synced != live:
            print("STALE: HQ's finance data does not match the workbook.\n"
                  "  run: python3 runtime/finance_model_sync.py", file=sys.stderr)
            return 1
        print("in sync")
        return 0

    data = extract()
    with open(OUT, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    t = data["scenarios"]["Target"]["years"][2]
    print(f"synced → dashboard/finance_model.json")
    print(f"  horizon {data['horizon']['start']} → {data['horizon']['end']}  ·  active: {data['activeScenario']}")
    print(f"  Target Y3: {t['clients']:.0f} clients · ${t['revenue']:,.0f} revenue · "
          f"${t['ebitda']:,.0f} EBITDA · NOI {t['noiPct']:.1%}")
    print(f"  sha256 {live[:12]}…")
    return 0


if __name__ == "__main__":
    sys.exit(main())
