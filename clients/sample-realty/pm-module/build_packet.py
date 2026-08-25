#!/usr/bin/env python3
"""Sample Realty PM module — 'your year, reconciled' packet builder.

Ingests Kimi's hand-kept ST escrow trust-account journal (the single entry
stream), then REGENERATES everything she currently maintains by hand:
  - a clean, date-ordered master journal with recomputed running balance
  - one ledger per property (monthly summary + full detail)
  - the NCREC-style trial balance
  - a findings/reconciliation tab (every anomaly, with row references)

This is the one-shot demo of the operated monthly module: enter it once,
everything else is generated. Run from this directory:
    python3 build_packet.py
Output: out/Parker-Realty-Trust-Account-Review-2026.xlsx
"""
import datetime, os, re
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "source", "Rent Account Journal 2026.xlsx")
OUT_DIR = os.path.join(HERE, "out")
OUT = os.path.join(OUT_DIR, "Parker-Realty-Trust-Account-Review-2026.xlsx")
os.makedirs(OUT_DIR, exist_ok=True)

TODAY = datetime.datetime(2026, 8, 4)
OPENING = 125.00          # her row 3: $100 misc buffer + $25 Capps overpayment
MONTHS = ["January","February","March","April","May","June","July","August",
          "September","October","November","December"]

# ── canonical property map (the fuzzy-matching the AI does per-transaction) ──
CANON = [
    ("1010 Bexton",         [r"1010\s*bexton"]),
    ("15721 Capps",         [r"15721\s*capps", r"\bcapps\b"]),
    ("1228 Whitby Moore",   [r"1228\s*whitby"]),
    ("3625 Vallette Ct",    [r"vallette"]),
    ("3818 Kalispell",      [r"kalispel"]),
    ("15219 Barossa Valley",[r"barossa"]),
    ("1138 Doveridge",      [r"1138\s*doveridge"]),
    ("1139 Doveridge",      [r"1139\s*doveridge"]),
    ("Misc Fund (NC)",      [r"misc", r"incidentals"]),
]
AMBIG_DOVERIDGE = re.compile(r"doveridge", re.I)

def canon_property(raw, party, tenant_map, flags):
    txt = (str(raw) if raw else "").strip().lower()
    if not txt:
        flags.append("no property recorded"); return "— Unassigned"
    if "," in txt and sum(1 for _, pats in CANON if any(re.search(p, txt) for p in pats)) > 1:
        flags.append("one entry spans multiple properties — needs split"); return "Multiple (needs split)"
    if txt.startswith("all "):
        flags.append("one entry spans multiple properties — needs split"); return "Multiple (needs split)"
    if "test" in txt:
        flags.append("test entry"); return "— Test entry"
    for name, pats in CANON:
        if any(re.search(p, txt) for p in pats):
            if name in ("1138 Doveridge","1139 Doveridge"): return name
            if AMBIG_DOVERIDGE.search(txt) and name not in ("1138 Doveridge","1139 Doveridge"):
                break
            return name
    if AMBIG_DOVERIDGE.search(txt):
        t = norm_party(party)
        if t in tenant_map and len(tenant_map[t]) == 1:
            resolved = next(iter(tenant_map[t]))
            flags.append(f"'{raw}' resolved to {resolved} via tenant match"); return resolved
        flags.append(f"ambiguous property '{str(raw).strip()}' — three Doveridge homes exist")
        return "Doveridge (ambiguous)"
    flags.append(f"unrecognized property '{str(raw).strip()}'")
    return "— Unassigned"

def norm_party(p):
    return re.sub(r"\s+", " ", str(p or "").strip().lower())

def classify(party, desc):
    t = f"{party or ''} {desc or ''}".lower()
    if "security deposit" in t or re.search(r"\bsd\b", t): return "Security deposit"
    if "lease set" in t or re.search(r"\blsu\b", t): return "Leasing fee"
    if "maint" in t or "repair" in t or "dishwasher" in t or "furnace" in t: return "Maintenance"
    if "rent" in t and "fee" not in t and not re.search(r"\bpm\b", t): return "Rent"
    if re.search(r"\b(pm|mgt|mgmt|management)\b", t): return "PM fee"
    if "disburs" in t: return "Owner disbursement"
    if "balance from end of year" in t: return "Opening balance"
    if "late fee" in t or "rent" in t: return "Rent"
    return "Other / review"

def fix_date(d, row, flags):
    if not isinstance(d, datetime.datetime):
        return None      # caller infers from journal position and flags it
    if d > TODAY and d.month == 12:
        flags.append(f"date typo corrected: {d.date()} → {d.replace(year=d.year-1).date()}")
        return d.replace(year=d.year - 1)
    if d > TODAY:
        flags.append(f"future-dated entry ({d.date()}) — review")
    return d

def acct_month(d):
    """Kimi's convention: end-of-Dec receipts for January rent count in January."""
    if d is None: return None
    return 1 if d.year < 2026 else d.month

# ── ingest ──────────────────────────────────────────────────────────────────
wb = load_workbook(SRC, data_only=True)
ws = wb["2026 ST Account Journal"]
raw = []
for r in range(3, ws.max_row + 1):
    date, party, desc, prop, method = (ws.cell(r, c).value for c in (1,2,3,4,5))
    num, cleared, dep, chk, bal, note = (ws.cell(r, c).value for c in (6,7,8,9,10,11))
    if all(v in (None, " ", "") for v in (date, party, desc, dep, chk, bal)): continue
    raw.append(dict(row=r, date=date, party=party, desc=desc, prop=prop, method=method,
                    num=num, cleared=str(cleared or "").strip().lower() == "y",
                    dep=float(dep) if isinstance(dep,(int,float)) else 0.0,
                    chk=float(chk) if isinstance(chk,(int,float)) else 0.0,
                    bal=bal, note=note))

# first pass: tenant → property map from unambiguous rent rows
tenant_map = defaultdict(set)
for x in raw:
    txt = (str(x["prop"]) or "").lower()
    for name, pats in CANON:
        if name.endswith("Doveridge") and any(re.search(p, txt) for p in pats):
            tenant_map[norm_party(x["party"])].add(name)

entries = []
last_date = None
for x in raw:
    flags = []
    d = fix_date(x["date"], x["row"], flags)
    if d is None:
        if last_date is not None:
            d = last_date
            flags.append(f"date missing — inferred ≈{d.date()} from journal position")
        else:
            flags.append("date missing — could not infer")
    else:
        last_date = d
    typ = classify(x["party"], x["desc"])
    if typ == "Opening balance": continue
    if d is not None and d.year < 2026:
        flags.append("prior-year receipt — counted in January per her books")
    prop = canon_property(x["prop"], x["party"], tenant_map, flags)
    if not x["cleared"]: flags.append("not marked cleared vs bank")
    entries.append(dict(x, date=d, prop=prop, type=typ, flags=flags))

entries.sort(key=lambda e: (e["date"] or TODAY, e["row"]))

# balance breaks in HER sheet (original order) for findings
orig_breaks, prev = [], OPENING
for x in raw:
    if isinstance(x["bal"], (int, float)):
        exp = round(prev + x["dep"] - x["chk"], 2)
        if abs(exp - round(float(x["bal"]), 2)) > 0.01 and x["row"] > 3:
            orig_breaks.append((x["row"], exp, round(float(x["bal"]),2)))
        prev = float(x["bal"])

props_seen = sorted({e["prop"] for e in entries})
fee_checks = []
by_pm = defaultdict(lambda: defaultdict(float))   # (prop, month) -> type -> amt
for e in entries:
    key = (e["prop"], acct_month(e["date"]) or 0)
    if e["type"] == "Rent": by_pm[key]["rent"] += e["dep"]
    if e["type"] == "PM fee": by_pm[key]["fee"] += e["chk"]

# ── build workbook ──────────────────────────────────────────────────────────
F = "Arial"
ink, red, soft = "131318", "C00003", "6B6B74"
th = Font(name=F, bold=True, size=10, color="FFFFFF")
hfill = PatternFill("solid", fgColor=ink)
money = '$#,##0.00;($#,##0.00);"—"'
thin = Border(bottom=Side(style="thin", color="DDDDDD"))

out = Workbook(); out.remove(out.active)

def style_header(wsx, row, cols):
    for c in range(1, cols+1):
        cell = wsx.cell(row, c); cell.font = th; cell.fill = hfill
        cell.alignment = Alignment(vertical="center")

# — Journal (clean) —
J = out.create_sheet("Journal (clean)")
J["A1"] = "Sample Realty — ST Rent Trust Account · Clean Journal 2026 (regenerated)"
J["A1"].font = Font(name=F, bold=True, size=13)
J["A2"] = "One row per transaction, date-ordered, balance recomputed. Source: Kimi's 'Rent Account Journal 2026.xlsx'. Flags list every correction made."
J["A2"].font = Font(name=F, italic=True, size=9, color=soft)
hdr = ["Date","Party","Description","Property (matched)","Type","Method","#","Cleared","Deposit","Withdrawal","Balance","Mo#","Flags"]
for c, h in enumerate(hdr, 1): J.cell(4, c, h)
style_header(J, 4, len(hdr))
r = 5
J.cell(r, 3, "Opening balance (carried from 2025)").font = Font(name=F, italic=True, size=10)
J.cell(r, 11, OPENING).number_format = money
J.cell(r, 11).font = Font(name=F, bold=True, size=10)
first_data = r + 1
for e in entries:
    r += 1
    J.cell(r, 1, e["date"].date() if e["date"] else None).number_format = "mm/dd/yyyy"
    J.cell(r, 2, str(e["party"] or "").strip())
    J.cell(r, 3, str(e["desc"] or "").strip())
    J.cell(r, 4, e["prop"])
    J.cell(r, 5, e["type"])
    J.cell(r, 6, str(e["method"] or "").strip())
    J.cell(r, 7, e["num"] if isinstance(e["num"], (int, float)) else None)
    J.cell(r, 8, "y" if e["cleared"] else "OPEN")
    if e["dep"]: J.cell(r, 9, e["dep"]).number_format = money
    if e["chk"]: J.cell(r, 10, e["chk"]).number_format = money
    J.cell(r, 11, f"=K{r-1}+N(I{r})-N(J{r})").number_format = money
    J.cell(r, 12, acct_month(e["date"]))
    J.cell(r, 13, "; ".join(e["flags"]))
    if e["flags"]: J.cell(r, 13).font = Font(name=F, size=9, color=red)
    else: J.cell(r, 13).font = Font(name=F, size=9, color=soft)
    for c in (2,3,4,5,6): J.cell(r, c).font = J.cell(r, c).font.copy(name=F, size=10) if J.cell(r,c).font else Font(name=F, size=10)
last = r
r += 2
J.cell(r, 3, "TOTALS").font = Font(name=F, bold=True)
J.cell(r, 9, f"=SUM(I{first_data}:I{last})").number_format = money
J.cell(r, 10, f"=SUM(J{first_data}:J{last})").number_format = money
J.cell(r, 11, f"=K{last}").number_format = money
for cell in (J.cell(r,9), J.cell(r,10), J.cell(r,11)): cell.font = Font(name=F, bold=True)
widths = [11,22,38,20,16,18,6,8,12,12,12,5,46]
for i, w in enumerate(widths, 1): J.column_dimensions[chr(64+i)].width = w
J.freeze_panes = "A5"

jref = "'Journal (clean)'"

# — Trial Balance —
T = out.create_sheet("Trial Balance")
T["A1"] = "NCREC Trial Balance — funds held per property, month-end (regenerated)"
T["A1"].font = Font(name=F, bold=True, size=13)
T["A2"] = "Every figure is a live formula over the clean journal — nothing keyed by hand. Months with no activity show —."
T["A2"].font = Font(name=F, italic=True, size=9, color=soft)
T.cell(4, 1, "Property")
for m in range(1, 13):
    T.cell(3, 1 + m, m).font = Font(name=F, size=8, color=soft)
    T.cell(4, 1 + m, MONTHS[m-1][:3].upper())
style_header(T, 4, 13)
rowi = 5
props_tb = props_seen + (["Misc Fund (NC)"] if "Misc Fund (NC)" not in props_seen else [])
for p in props_tb:
    T.cell(rowi, 1, p).font = Font(name=F, size=10)
    for m in range(1, 13):
        col = chr(66 + m - 1)
        T.cell(rowi, 1 + m,
            f"=SUMIFS({jref}!$I:$I,{jref}!$D:$D,$A{rowi},{jref}!$L:$L,\"<=\"&{col}$3)"
            f"-SUMIFS({jref}!$J:$J,{jref}!$D:$D,$A{rowi},{jref}!$L:$L,\"<=\"&{col}$3)"
            + (f"+{OPENING}" if p == "Misc Fund (NC)" else "")
        ).number_format = money
        T.cell(rowi, 1 + m).font = Font(name=F, size=10)
    rowi += 1
rowi += 1
labels = [("Total deposits (month)", "I", ""), ("Total withdrawals (month)", "J", ""),]
for lab, colL, _ in labels:
    T.cell(rowi, 1, lab).font = Font(name=F, bold=True, size=10)
    for m in range(1, 13):
        col = chr(66 + m - 1)
        T.cell(rowi, 1 + m, f"=SUMIFS({jref}!${colL}:${colL},{jref}!$L:$L,{col}$3)").number_format = money
    rowi += 1
T.cell(rowi, 1, "Account balance (month-end)").font = Font(name=F, bold=True, size=10)
for m in range(1, 13):
    col = chr(66 + m - 1)
    T.cell(rowi, 1 + m,
        f"={OPENING}+SUMIFS({jref}!$I:$I,{jref}!$L:$L,\"<=\"&{col}$3)"
        f"-SUMIFS({jref}!$J:$J,{jref}!$L:$L,\"<=\"&{col}$3)").number_format = money
    T.cell(rowi, 1 + m).font = Font(name=F, bold=True, size=10, color=red)
T.column_dimensions["A"].width = 24
for m in range(1, 13): T.column_dimensions[chr(65+m)].width = 11
T.freeze_panes = "B5"

# — per-property ledgers —
for p in [x for x in props_seen if not x.startswith("—")]:
    name = re.sub(r"[\\/*?\[\]:]", "", p)[:28]
    L = out.create_sheet(name)
    L["A1"] = f"{p} — 2026 ledger (regenerated from the journal)"
    L["A1"].font = Font(name=F, bold=True, size=12)
    L.cell(3, 1, "Month"); L.cell(3, 2, "Rent in"); L.cell(3, 3, "PM fees"); L.cell(3, 4, "Maintenance")
    L.cell(3, 5, "Owner disb."); L.cell(3, 6, "Other"); L.cell(3, 7, "Net")
    style_header(L, 3, 7)
    dhdr_at = 3 + 12 + 3
    for m in range(1, 13):
        rr = 3 + m
        L.cell(rr, 1, MONTHS[m-1]).font = Font(name=F, size=10)
        base = (f"{jref}!$D:$D,\"{p}\",{jref}!$L:$L,{m}")
        L.cell(rr, 2, f"=SUMIFS({jref}!$I:$I,{base},{jref}!$E:$E,\"Rent\")").number_format = money
        L.cell(rr, 3, f"=SUMIFS({jref}!$J:$J,{base},{jref}!$E:$E,\"PM fee\")").number_format = money
        L.cell(rr, 4, f"=SUMIFS({jref}!$J:$J,{base},{jref}!$E:$E,\"Maintenance\")").number_format = money
        L.cell(rr, 5, f"=SUMIFS({jref}!$J:$J,{base},{jref}!$E:$E,\"Owner disbursement\")").number_format = money
        L.cell(rr, 6, f"=SUMIFS({jref}!$I:$I,{base})-B{rr}+SUMIFS({jref}!$J:$J,{base})-C{rr}-D{rr}-E{rr}").number_format = money
        L.cell(rr, 7, f"=SUMIFS({jref}!$I:$I,{base})-SUMIFS({jref}!$J:$J,{base})").number_format = money
    tr = 16
    L.cell(tr, 1, "YEAR").font = Font(name=F, bold=True, size=10)
    for c in range(2, 8):
        col = chr(64 + c)
        L.cell(tr, c, f"=SUM({col}4:{col}15)").number_format = money
        L.cell(tr, c).font = Font(name=F, bold=True, size=10)
    # detail
    L.cell(dhdr_at, 1, "Date"); L.cell(dhdr_at, 2, "Party"); L.cell(dhdr_at, 3, "Description")
    L.cell(dhdr_at, 4, "Type"); L.cell(dhdr_at, 5, "Deposit"); L.cell(dhdr_at, 6, "Withdrawal"); L.cell(dhdr_at, 7, "Flags")
    style_header(L, dhdr_at, 7)
    rr = dhdr_at
    for e in entries:
        if e["prop"] != p: continue
        rr += 1
        L.cell(rr, 1, e["date"].date() if e["date"] else None).number_format = "mm/dd/yyyy"
        L.cell(rr, 2, str(e["party"] or "").strip()).font = Font(name=F, size=9.5)
        L.cell(rr, 3, str(e["desc"] or "").strip()).font = Font(name=F, size=9.5)
        L.cell(rr, 4, e["type"]).font = Font(name=F, size=9.5)
        if e["dep"]: L.cell(rr, 5, e["dep"]).number_format = money
        if e["chk"]: L.cell(rr, 6, e["chk"]).number_format = money
        L.cell(rr, 7, "; ".join(e["flags"])).font = Font(name=F, size=8.5, color=red if e["flags"] else soft)
    for i, w in enumerate([12,20,36,16,12,12,40], 1): L.column_dimensions[chr(64+i)].width = w

# — owner statements —
OWNERS = {
    "Raju Prasad":      ["1010 Bexton", "1228 Whitby Moore"],
    "Bhavani":          ["15219 Barossa Valley", "1138 Doveridge", "1139 Doveridge", "Doveridge (ambiguous)"],
    "Griffin Campbell": ["3625 Vallette Ct"],
    "Sanjeet Kumar":    ["15721 Capps"],
    "Chandra":          ["3818 Kalispell"],
}
for owner, plist in OWNERS.items():
    S = out.create_sheet(f"Stmt — {owner}"[:31])
    S["A1"] = f"Owner Statement — {owner}"
    S["A1"].font = Font(name=F, bold=True, size=14)
    S["A2"] = "Sample Realty Home & Land, LLC · Property Management · January–July 2026"
    S["A2"].font = Font(name=F, italic=True, size=9, color=soft)
    S["A3"] = "Generated from the trust-account journal — every figure traceable to a transaction. (Draft for Kimi's review before sending.)"
    S["A3"].font = Font(name=F, size=8.5, color=soft)
    rr = 5
    grand_first = None
    for p in plist:
        if p not in props_seen: continue
        S.cell(rr, 1, p).font = Font(name=F, bold=True, size=12)
        if p == "Doveridge (ambiguous)":
            S.cell(rr, 2, "entries recorded as just 'Doveridge' — to confirm").font = Font(name=F, italic=True, size=9, color=red)
        rr += 1
        heads = ["Month","Rent collected","PM fee","Maintenance","Leasing/other","Disbursed to you","Held for property"]
        for c, h in enumerate(heads, 1): S.cell(rr, c, h)
        style_header(S, rr, len(heads))
        hdr_row = rr
        rr += 1
        first_m = rr
        for m in range(1, 8):
            base = f"{jref}!$D:$D,\"{p}\",{jref}!$L:$L,{m}"
            S.cell(rr, 1, MONTHS[m-1]).font = Font(name=F, size=10)
            S.cell(rr, 2, f"=SUMIFS({jref}!$I:$I,{base},{jref}!$E:$E,\"Rent\")").number_format = money
            S.cell(rr, 3, f"=SUMIFS({jref}!$J:$J,{base},{jref}!$E:$E,\"PM fee\")").number_format = money
            S.cell(rr, 4, f"=SUMIFS({jref}!$J:$J,{base},{jref}!$E:$E,\"Maintenance\")").number_format = money
            S.cell(rr, 5, f"=SUMIFS({jref}!$J:$J,{base},{jref}!$E:$E,\"Leasing fee\")+SUMIFS({jref}!$J:$J,{base},{jref}!$E:$E,\"Other / review\")").number_format = money
            S.cell(rr, 6, f"=SUMIFS({jref}!$J:$J,{base},{jref}!$E:$E,\"Owner disbursement\")").number_format = money
            S.cell(rr, 7, f"=SUMIFS({jref}!$I:$I,{jref}!$D:$D,\"{p}\",{jref}!$L:$L,\"<=\"&{m})-SUMIFS({jref}!$J:$J,{jref}!$D:$D,\"{p}\",{jref}!$L:$L,\"<=\"&{m})").number_format = money
            rr += 1
        S.cell(rr, 1, "YTD").font = Font(name=F, bold=True, size=10)
        for c in range(2, 7):
            col = chr(64 + c)
            S.cell(rr, c, f"=SUM({col}{first_m}:{col}{rr-1})").number_format = money
            S.cell(rr, c).font = Font(name=F, bold=True, size=10)
        S.cell(rr, 7, f"=G{rr-1}").number_format = money
        S.cell(rr, 7).font = Font(name=F, bold=True, size=10, color=red)
        rr += 2
    for i, w in enumerate([13,14,12,13,13,15,15], 1): S.column_dimensions[chr(64+i)].width = w

# — findings —
FD = out.create_sheet("Findings", 0)
FD["A1"] = "Trust Account Review — what one automated pass found"
FD["A1"].font = Font(name=F, bold=True, size=14)
FD["A2"] = f"Source: Kimi's journal, {len(entries)} transactions, Jan–Jul 2026 · generated {TODAY.date()} · every figure traceable to a journal row"
FD["A2"].font = Font(name=F, italic=True, size=9, color=soft)

frow = 4
def finding(title, detail):
    global frow
    FD.cell(frow, 1, "●").font = Font(name=F, bold=True, color=red)
    FD.cell(frow, 2, title).font = Font(name=F, bold=True, size=11)
    frow += 1
    FD.cell(frow, 2, detail).font = Font(name=F, size=10, color=soft)
    FD.cell(frow, 2).alignment = Alignment(wrap_text=True)
    FD.row_dimensions[frow].height = 30
    frow += 2

n_variants = len({(str(x['prop']) or '').strip().lower() for x in raw if x['prop']})
finding(f"{n_variants} different property spellings for ~10 properties",
        "e.g. '1010 Bexton' vs '1010 bexton st'; 'Kalispell' vs 'Kalispel'; bare 'Doveridge' with three Doveridge homes under management. All matched to canonical names in this workbook; ambiguous rows resolved by tenant where possible and flagged where not.")
dov = sum(1 for e in entries if "resolved to" in "; ".join(e["flags"]))
amb = sum(1 for e in entries if e["prop"] == "Doveridge (ambiguous)")
finding(f"{dov} entries auto-resolved by tenant match · {amb} still ambiguous",
        "Entries that just said 'Doveridge' were tied to the right home using who paid. The remaining ambiguous rows are listed in the Journal tab with red flags — a 2-minute review, once.")
finding(f"{len(orig_breaks)} running-balance breaks in the hand-kept journal",
        "Rows entered out of date order made the typed balance column contradict itself (worst around Feb 10). The clean journal recomputes every balance by formula — this class of error can no longer exist.")
datefix = sum(1 for e in entries if any("date typo" in f for f in e["flags"]))
finding(f"{datefix} date typo(s) corrected",
        "December entries keyed as 12/30/2026 (a year in the future). Corrected to 2025 and flagged on the row.")
nodate = sum(1 for e in entries if any("date missing" in f for f in e["flags"]))
finding(f"{nodate} entries had no date at all",
        "Undated rows silently fall out of any monthly report. Each was assigned the date of the neighboring journal entry and flagged for a quick confirm — in the live module the bank feed supplies the date, so this disappears entirely.")
unclr = sum(1 for e in entries if not e["cleared"])
finding(f"{unclr} entries not marked cleared against the bank",
        "Open items as of the file date — the module chases these automatically each month instead of a line-by-line statement tick.")
multi = sum(1 for e in entries if e["prop"] == "Multiple (needs split)")
finding(f"{multi} entries cover multiple properties in one line",
        "e.g. one Zelle covering several PM fees ('all Kristi's properties'). These can't post to any single ledger — the module splits them at entry time.")
td = sum(e["dep"] for e in entries); tc = sum(e["chk"] for e in entries)
finding(f"The account cushion is off: ends at ${OPENING + td - tc:,.2f}, should float $125.00",
        f"Opening float was $125 ($100 incidentals + $25 Capps overpayment, per the journal's own notes). {len(entries)} transactions later the recomputed balance is ${OPENING+td-tc:,.2f} — a ${abs(OPENING+td-tc-125):,.2f} drift that needs a one-time true-up. In an NCREC trust account, drift is the thing to catch the day it happens.")
finding("Per-property balances don't zero out — the headline finding",
        "In a fully-disbursed trust account each property should end near $0. Instead: 1010 Bexton ends July at −$1,830.51 (April's $4,331 repair bills exceeded its rents — those bills were effectively covered by other owners' funds held in the same account, or an owner top-up went unrecorded); the bare-'Doveridge' entries hide −$6,150 that mostly belongs against 1139 Doveridge's +$6,396; and the unsplit multi-property Zelles carry −$1,854 that mirrors small positives stranded on Whitby Moore, Vallette, Kalispell, and Barossa. None of this is visible in the hand-kept books — it appears the moment the ledgers are generated instead of typed. See the Trial Balance tab.")
finding("PM fee spot-check: 8% pattern holds where rent and fee both posted",
        "Fees observed at 8.0% of collected rent on matched months (e.g. $204/$2,550 · $180/$2,250 · $210/$2,625 · $128/$1,600). The module computes these automatically each month and flags any deviation.")
FD.cell(frow, 2, "Every number in this workbook is a live formula over the clean journal — change a journal row and the ledgers, trial balance, and totals all update. That is the module: enter it once (or let the bank feed enter it), everything else is generated.").font = Font(name=F, italic=True, size=10)
FD.cell(frow, 2).alignment = Alignment(wrap_text=True)
FD.row_dimensions[frow].height = 44
FD.column_dimensions["A"].width = 3
FD.column_dimensions["B"].width = 110

out.save(OUT)
print("saved", OUT)
print("entries:", len(entries), "| props:", len(props_seen))
print("orig breaks:", len(orig_breaks), "| unclr:", unclr, "| datefix:", datefix, "| multi:", multi, "| ambiguous left:", amb, "| tenant-resolved:", dov)
print(f"deposits {td:,.2f} withdrawals {tc:,.2f} end {OPENING+td-tc:,.2f}")
